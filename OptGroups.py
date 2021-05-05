from __future__ import division

from openpyxl import Workbook
from openpyxl.styles import Font, Color

import copy
import numpy
import random


## Global Variables:


MEMBERS = [                 ###   MEMBERS - list of the members which contain name,average and sex.
  ["Adam", "5", "M"],       ###   this list gets the input from the User costumized table.
  ["Ali", "2", "F"],        ###   the sex variable need to get "M" for Male or "F" for Female.
  ["Andrew", "1", "M"],
  ["Blonde Megan", "5", "F"],
  ["Brian", "3", "M"],
  ["Brian Tall", "4", "F"],
  ["Collin", "4", "M"],
  ["Dave", "2.5", "M"],
  ["Dylan", "4.8", "F"],
  ["Jake", "3.6", "F"],
  ["James", "4.5", "M"],
  ["Jason", "3.9", "F"],
  ["Jdao", "2.6", "F"],
  ["Jimbo", "4.2", "M"],
  ["Joe", "4.6", "M"],
  ["John", "3.9", "M"],
  ["Lauren", "3.5", "F"],
  ["Mark", "5", "M"],
  ["Matt", "4.3", "F"],
  ["Meaghan Creamer", "2.6", "F"],
  ["Ncik", "3.7", "M"],
  ["Nick", "4.3", "M"],
  ["Nicole", "4.1", "F"],
  ["Olivia", "5", "F"],
  ["Parks", "4.7", "M"],
  ["Poot", "2.1", "M"],
  ["Sam Tall", "3.2", "F"],
  ["Shaundry", "4.8", "F"],
  ["Tommy Doug", "4.6", "M"],
  ["Tucker", "4.3", "M"]

]

workbookName = "Groups.xlsx"                                                                  ##  the name of the output Excel file

###   NUM_PARTICIPANTS and NUM_TEAMS are variables which one of them is required.  ###
###   user needs to choose one of these options to determine how the groups will split. ###

# NUM_PARTICIPANTS = input("How many participants would you like the groups to contain? ")      ##  How many groups required

# if NUM_PARTICIPANTS=="":
#   NUM_TEAMS =int( input("How many gruops would you like to split to? "))                      ##  How many groups required
# else :
#   NUM_TEAMS= round(len(MEMBERS)/int(NUM_PARTICIPANTS))                                        ##  How many groups required if we get participants per group

# NUM_RUNS = 5000         ## how many iterations ( Heuristic algorithm)

# Statistic weights - explanation found in Solution class
# avg of input
AVERAGE= 1
TEAM_SIZE_WEIGHT = 10
GENDER_WEIGHT = 100


class Player(object):
  def __init__(self, name, average, gender):
    self.name = name
    self.average = average
    self.gender = gender
    self.listPlayer = [name,gender]


  # @staticmethod
  # def _get_numeric_time(time):
  #   """time is of the form "H:MM:SS". We convert to total seconds"""
  #   _, minutes, seconds = map(int, re.split(":", time))
  #   return minutes * 60 + seconds

  @staticmethod
  def _average_toInt(average):
    return round(float(average),2)

  def get_average(self):
    return self._average_toInt(self.average)

  def is_male(self):
    return self.gender == "Male"


class Team(set):
  def __str__(self):
    self.teamList=[player.listPlayer for player in self]
    return("  Average: {}. {} players, {} male: {}".format(
      self.total_average(), len(self), self.num_males(),
      [player.listPlayer for player in self]))

  def total_average(self):                                  ## Average calculate for current group
    return sum([player.get_average() for player in self])

  def num_males(self):
    return len([player for player in self if player.is_male()])



class Solution(object):
  def __init__(self):
    self.teams = []
    for _ in range(NUM_TEAMS):
      self.teams.append(Team())

  def group_score(self):
    # We use the measures of three variances to determine our score:
    #  - total team mile time
    #  - num males on the team
    #  - team size
    #
    # The lowest score will be the one that does the best at minimizing the
    # difference between teams in these catagories.
    average_variance = numpy.var([team.total_average() for team in self.teams])
    team_size_variance = numpy.var([len(team) for team in self.teams])
    gender_variance = numpy.var([team.num_males() for team in self.teams])
    return(average_variance * AVERAGE + team_size_variance * TEAM_SIZE_WEIGHT
           + gender_variance * GENDER_WEIGHT)

  def add_player_to_random_team(self, player):
    random.choice(self.teams).add(player)

  def change_random_player(self):
    old_team = random.choice(self.teams)
    if len(old_team) == 0:
      print ("empty team")
      return
    player = random.sample(old_team, 1)[0]
    old_team.remove(player)
    new_team = random.choice(self.teams)
    new_team.add(player)

  def __str__(self):
    return "\n".join(map(str, self.teams))

  def exportXl(self):     ## create excel file with the solution divided to worksheet per gruop
      wb = Workbook()
      for i in range(1,(NUM_TEAMS+1)):
          ws = wb.create_sheet("Group {}".format(i))
          ws.merge_cells('A1:B1')
          ws['A1'].value = ('Group {}'.format(i))
          ws['A1'].font = Font(color='DC143C', italic=True)
          for row in self.teams[i-1]:
              ws.append(row.listPlayer)
      del (wb['Sheet'])
      wb.save(filename = workbookName)



def main():
  initial_solution = Solution()

  for name, average, gender in MEMBERS:           ## create random groups for the begining
    initial_solution.add_player_to_random_team(Player(name, average, gender))

  print("Splitting {} players ({} male) into {} teams".format(            ## print starting information
    len(MEMBERS), len([runner for runner in MEMBERS if runner[2] == "Yes"]),
    NUM_TEAMS))

  print("Starting with solution score {:.2f}:\n{}".format(        ## print the non balanced groups
    initial_solution.group_score(), initial_solution))

  for run_num in range(NUM_RUNS):       ## creating groups

    if run_num % 1000 == 0:

      print("Current best solution with score {:.2f}:\n{}".format(    ## print the current score
        initial_solution.group_score(), initial_solution))

    solution = copy.deepcopy(initial_solution)                        ## copy the current solution object into new variable to compare after

    for _ in range(random.randint(1, 10)):
      solution.change_random_player()
    if solution.group_score() < initial_solution.group_score():
      initial_solution = solution

  print("Best solution found, with solution score {:.2f}:\n{}".format(
    initial_solution.group_score(), initial_solution))
  print('\n\n')

  initial_solution.exportXl()       ## output the optimal solution to excel file


# if __name__ == "__main__":      ## START THE PROGRAM
#   main()


