"""This is the main flask python file"""

from smtplib import SMTPAuthenticationError
from MySQLdb._exceptions import IntegrityError
from flask import Flask, request, render_template, redirect, url_for, session, g
import seed
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import os
import numpy as np
import copy
import random
from OptGroups import Solution
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from flask_mysqldb import MySQL
from flask_mail import Mail, Message
from itsdangerous import URLSafeTimedSerializer, SignatureExpired

NUM_TEAMS = None
NUM_RUNS = 5000
AVERAGE = 1
TEAM_SIZE_WEIGHT = 10
GENDER_WEIGHT = 100
workbookName = "Groups.xlsx"


class Player(object):
    def __init__(self, name, average, gender):
        self.name = name
        self.average = average
        self.gender = gender
        self.listPlayer = [name, gender]

    # @staticmethod
    # def _get_numeric_time(time):
    #   """time is of the form "H:MM:SS". We convert to total seconds"""
    #   _, minutes, seconds = map(int, re.split(":", time))
    #   return minutes * 60 + seconds

    @staticmethod
    def _average_toInt(average):
        return round(float(average), 2)

    def get_average(self):
        return self._average_toInt(self.average)

    def is_male(self):
        return self.gender == "Yes"


class Team(set):
    def __str__(self):
        self.teamList = [player.listPlayer for player in self]
        return "  Average: {}. {} players, {} male: {}".format(
            self.total_average(),
            len(self),
            self.num_males(),
            [player.listPlayer for player in self],
        )

    def total_average(self):  # Average calculate for current group
        return sum([player.get_average() for player in self])

    def num_males(self):
        return len([player for player in self if player.is_male()])


class Solution(object):
    def __init__(self):
        self.teams = []
        for _ in range(NUM_TEAMS):
            self.teams.append(Team())

    def group_score(self):
        # We use the measures of three variances to determine our score:
        #  - total team mile time
        #  - num males on the team
        #  - team size
        #
        # The lowest score will be the one that does the best at minimizing the
        # difference between teams in these catagories.
        average_variance = np.var([team.total_average()
                                   for team in self.teams])
        team_size_variance = np.var([len(team) for team in self.teams])
        gender_variance = np.var([team.num_males() for team in self.teams])
        return (
            average_variance * AVERAGE
            + team_size_variance * TEAM_SIZE_WEIGHT
            + gender_variance * GENDER_WEIGHT
        )

    def add_player_to_random_team(self, player):
        random.choice(self.teams).add(player)

    def change_random_player(self):
        old_team = random.choice(self.teams)
        if len(old_team) == 0:
            print("empty team")
            return
        player = random.sample(old_team, 1)[0]
        old_team.remove(player)
        new_team = random.choice(self.teams)
        new_team.add(player)

    def __str__(self):
        return "\n".join(map(str, self.teams))

    def exportXl(
            self,
    ):  # create excel file with the solution divided to worksheet per group
        wb = Workbook()
        excel_dict = {}
        for i in range(1, (NUM_TEAMS + 1)):
            excel_dict[i] = {}
            ws = wb.create_sheet("Group {}".format(i))
            ws.merge_cells("A1:B1")

            ws["A1"].value = "Group {}".format(i)

            ws["A1"].font = Font(color="DC143C", italic=True)
            for j, row in enumerate(self.teams[i - 1]):
                excel_dict[i][j] = row.listPlayer
                ws.append(row.listPlayer)

        del wb["Sheet"]
        filepath = os.path.join("./static/output", workbookName)
        wb.save(filename=filepath)
        return excel_dict


class OptionArray(object):
    def __init__(self):
        self.option = 0
        self.excel_dict = {}

    def setExcel_Dict(self, dict, opt):
        self.excel_dict = dict
        self.option = opt


class ErrorMessages(object):
    def __init__(self):
        self.msg = ""
        self.id = 0

    def confirmationMessage(self,):
        self.msg = "A confirmation link has been sent to your email"
        self.id = 1

    def emailExist(self,):
        self.msg = "Email already exists."
        self.id = 2

    def serverBusy(self,):
        self.msg = "Server is busy. Try Again Later"
        self.id = 3

    def emailIncorrect(self,):
        self.msg = "Email is incorrect."
        self.id = 4

    def passwordIncorrect(self,):
        self.msg = "Password is incorrect."
        self.id = 5

    def emailConfirmed(self,):
        self.msg = "Your email is confirmed."
        self.id = 6

    def tokenExpired(self):
        self.msg = "Your token is expired."
        self.id = 7

    def emailDoesntExist(self):
        self.msg = "Email doesn't exist."
        self.id = 8

    def checkEmail(self):
        self.msg = "Check your email for further process."
        self.id = 9

    def passwordChanged(self):
        self.msg = "Your password is changed."
        self.id = 10

    def passDoesntMatched(self):
        self.msg = "Password doesn't matched"
        self.id = 11

    def passwordDoesntChanged(self):
        self.msg = "Password doesn't changed"
        self.id = 12

    def accessDenied(self):
        self.msg = "ACCESS DENIED"
        self.id = 13

    def profileUpdated(self):
        self.msg = "Profile Updated"
        self.id = 14


app = Flask(__name__)

# Database Details
app.config['MYSQL_HOST'] = seed.host
app.config['MYSQL_USER'] = seed.user
app.config['MYSQL_PASSWORD'] = seed.password
app.config['MYSQL_PORT'] = seed.port
app.config['MYSQL_DB'] = seed.db_name


# EMAIL SMTP DETAILS
app.config['MAIL_SERVER'] = "smtp.gmail.com"
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USE_SSL'] = False
app.config['MAIL_USERNAME'] = "legendbest123@gmail.com"
app.config['MAIL_PASSWORD'] = "gcpvxzooaxjqabdb"
app.config['MAIL_DEFAULT_SENDER'] = "legendbest123@gmail.com"
app.config['MAIL_ASCII_ATTACHMENTS'] = False

# SECRET KEY

app.secret_key = seed.secret_key
token_key = URLSafeTimedSerializer(seed.secret_key)


mysql = MySQL(app)
mail = Mail(app)

# FOR HOME PAGE #
####################################
# BACKEND: COMPLETED
# FRONTEND:
####################################


@app.route("/")
def index():
    if 'email' in session:
        g.email = session['email']
        return redirect(url_for('guidelines'))
    return render_template("index.html")


@app.after_request
def after_request(response):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return response
# FOR SIGN UP AND LOGIN PAGE ##

####################################
# BACKEND: COMPLETED
# FRONTEND: COMPLETED
####################################


@app.route("/signup", methods=['GET', 'POST'])
def signup():
    if 'email' in session:
        g.email = session['email']
        return redirect(url_for('guidelines'))
    else:
        if request.method == 'GET':
            return render_template('signup.html')

        else:
            if request.form['action'] == "Signup":
                email = request.form['email']
                fname = request.form['fname']
                lname = request.form['lname']
                password = request.form['password']
                organization = request.form['organization']

                hashedPassword = generate_password_hash(password)
                errormsg = ErrorMessages()
                try:
                    cur = mysql.connection.cursor()
                    cur.execute("INSERT INTO users(fname,lname,email,password,organization) VALUES (%s,%s,%s,%s,%s)",
                                (fname, lname, email, hashedPassword, organization))
                    token = token_key.dumps(email, salt='email-confirmation')
                    msg = Message('Email Confirmation from opTeam',
                                  sender="legendbest123@gmail.com", recipients=[email])
                    link = url_for('confirmEmail', token=token, _external=True)
                    msg.body = 'Your confirmation link is : {}'.format(link)
                    mail.send(msg)
                    # Token is generated sent to the email for confirmation
                    errormsg.confirmationMessage()
                    mysql.connection.commit()
                    cur.close()
                    return render_template('signup.html', signup=errormsg)
                except IntegrityError:
                    errormsg.emailExist()
                except SMTPAuthenticationError:
                    errormsg.serverBusy()
                except:
                    errormsg.emailExist()
                return render_template('signup.html', signup=errormsg)

            elif request.form['action'] == "Login":
                errormsg = ErrorMessages()
                session.pop('email', None)
                session.pop('name', None)
                email = request.form['email-login']
                password = request.form['password-login']
                try:
                    cur = mysql.connection.cursor()
                    cur.execute(
                        '''SELECT fname,lname,email,password,organization FROM users WHERE email=%s and confirmation=%s''', (email, 1))
                    person = cur.fetchone()
                    if person is None:
                        errormsg.emailIncorrect()
                    else:
                        if check_password_hash(person[3], password):
                            session['name'] = person[0] + ' ' + person[1]
                            session['email'] = person[2]
                            session['fname'] = person[0]
                            session['lname'] = person[1]
                            session['organization'] = person[4]
                            mysql.connection.commit()
                            return redirect(url_for('guidelines'))
                        else:
                            errormsg.passwordIncorrect()
                except:
                    errormsg.serverBusy()
        return render_template('signup.html', login=errormsg)

# FOR CONFIRM EMAIL #


@app.after_request
def after_request(response):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return response


@app.route('/confirm-email<token>')
def confirmEmail(token):

    errormsg = ErrorMessages()
    try:
        email = token_key.loads(token, salt='email-confirmation', max_age=3600)
        cur = mysql.connection.cursor()
        cur.execute("UPDATE users SET confirmation = %s WHERE email = %s",
                    (1, email))

        msg = Message('Email Confirmation from opTeam',
                      sender="legendbest123@gmail.com", recipients=[email])
        msg.body = 'Your email is confirmed.\n Thanks you for using OPTEAM'
        mail.send(msg)

        errormsg.emailConfirmed()
        mysql.connection.commit()
        cur.close()
    except SignatureExpired:
        errormsg.tokenExpired()

    return render_template('signup.html', signup=errormsg)


@app.after_request
def after_request(response):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return response
# FOR FORGET PASSWORD #


"""
Requirement: Functional Requirement
Type: Compulsory
"""
####################################
# BACKEND:  COMPLETED
# FRONTEND: COMPLETED
####################################


@app.route("/forget-password")
def forget_password():
    if 'email' in session:
        g.email = session['email']
        return redirect(url_for('guidelines'))
    return render_template("forget-password.html")


@app.after_request
def after_request(response):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return response


@app.route("/change-password", methods=['POST'])
def changePassword():
    errormsg = ErrorMessages()
    if request.method == 'POST':
        if request.form['action'] == 'forget-password':
            email = request.form['email']
            # print(email)
            try:
                cur = mysql.connection.cursor()
                cur.execute("SELECT email,confirmation FROM users WHERE email=%s",
                            (email,))

                person = cur.fetchone()
                if person is None:
                    errormsg.emailDoesntExist()
                if (person[1] == 0):
                    errormsg.emailDoesntExist()
                else:
                    token = token_key.dumps(person[0], salt='changepassword')
                    msg = Message('Forget Password from OPTEAM',
                                  sender="legendbest123@gmail.com", recipients=[person[0]])
                    link = url_for('password_confirmation',
                                   token=token, _external=True)
                    msg.body = 'Go to the link to change password : {}'.format(
                        link)
                    mail.send(msg)
                    mysql.connection.commit()
                    # Token is generated sent to the email for confirmation
                    errormsg.checkEmail()
            except SMTPAuthenticationError:
                errormsg.serverBusy()
            except:
                errormsg.emailDoesntExist()
            return render_template('forget-password.html', passwordMsg=errormsg)


@app.after_request
def after_request(response):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return response


@app.route("/change-password<token>")
def password_confirmation(token):
    email = token_key.loads(token, salt="changepassword", max_age=3600)
    return render_template('update-password.html', email=email)


@app.after_request
def after_request(response):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return response


@app.route("/new-password-update", methods=['POST'])
def new_password():
    errormsg = ErrorMessages()
    if (request.method == 'POST'):
        email = request.form['email']
        password = request.form['pass']
        confirmPass = request.form['confirmpass']
        if (password == confirmPass):
            hashedPassword = generate_password_hash(password)
            try:
                cur = mysql.connection.cursor()
                cur.execute("UPDATE users SET password = %s WHERE email = %s",
                            (hashedPassword, email))
                errormsg.passwordChanged()
                mysql.connection.commit()
                cur.close()
                return render_template('signup.html', login=errormsg)
            except SignatureExpired:
                errormsg.passwordDoesntChanged()
        errormsg.passDoesntMatched()
        return render_template('update-password.html', msg=errormsg)


@app.after_request
def after_request(response):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return response
# FOR RESULT PAGE #

####################################
# BACKEND: COMPLETED
# FRONTEND: COMPLETED
####################################


@app.route("/completed")
def complete():
    errormsg = ErrorMessages()
    if 'email' in session:
        g.email = session['email']
        return render_template("complete.html")
    errormsg.accessDenied()
    return render_template("signup.html", access=errormsg)


def after_request(response):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return response

# FOR GUIDELINES #

####################################
# BACKEND: COMPLETED
# FRONTEND: COMPLETED
####################################


@app.route("/guidelines")
def guidelines():
    errormsg = ErrorMessages()
    if 'email' in session:
        g.email = session['email']
        return render_template('guidelines.html')
    errormsg.accessDenied()
    return render_template("signup.html", access=errormsg)


def after_request(response):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return response

# FOR GROUP INFORMATION #

# MAIN APP1 PAGE
####################################
# BACKEND: COMPLETED
# FRONTEND: COMPLETED
####################################


@app.route("/group-information")
def groupInfo():
    errormsg = ErrorMessages()
    if 'email' in session:
        g.email = session['email']
        return render_template("group-information.html")
    errormsg.accessDenied()
    return render_template("signup.html", access=errormsg)


@app.after_request
def after_request(response):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return response
#####################################################

# FOR LOGOUT #
####################################
# BACKEND: COMPLETED
####################################
# LOGOUT IS DONE COMPLETELY


@app.route("/logout")
def logout():
    errormsg = ErrorMessages()
    if 'email' in session:
        session.pop('email', None)
        session.pop('name', None)
        return redirect(url_for('index'))
    else:
        errormsg.accessDenied()
        return render_template("signup.html", access=errormsg)


@app.after_request
def after_request(response):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return response

#####################################################
####################################
# BACKEND: COMPLETED
# FRONTEND: COMPLETED
####################################


@app.route("/process", methods=["POST"])
def process():
    if 'email' in session:
        global NUM_TEAMS
        members = []
        option = 0
        if request.method == "POST":

            if request.form['action'] == 'generate':
                option = 1
                # storing size of form values like number of participants
                rows = int((len(request.form) - 1) / 5)
                print(rows)
                for i in range(1, rows + 1):
                    print(i)
                    row_array = []
                    row_array.append(request.form["name{}".format(i)])
                    avg_integers = (
                        int(request.form["integer1_{}".format(i)])
                        + int(request.form["integer2_{}".format(i)])
                        + int(request.form["integer3_{}".format(i)])
                    ) / 3

                    row_array.append(str(avg_integers))
                    row_array.append(request.form["binary{}".format(i)])
                    members.append(row_array)
                # for storing data into array
                if "Number of Participants" in request.form:
                    NUM_PARTICIPANTS = int(
                        request.form["Number of Participants"])
                    NUM_TEAMS = round(len(members) / int(NUM_PARTICIPANTS))
                else:
                    NUM_TEAMS = int(request.form["groups"])
                initial_solution = Solution()

            elif request.form['action'] == 'excelButton':
                option = 2
                file = request.files["input-file"]
                foo = file.filename
                filepath = os.path.join('static', foo)
                file.save(filepath)
                excel_file = load_workbook(filepath)
                sheet_obj = excel_file.active
                count = 0

                for row in sheet_obj:

                    if count < 5:
                        count += 1

                    else:
                        one_member = []
                        one_member.append(row[0].value)
                        avg_integers = (
                            int(row[2].value) + int(row[3].value) + int(row[4].value)) / 3
                        one_member.append(str(avg_integers))
                        one_member.append(row[1].value)
                        members.append(one_member)
                os.remove(filepath)
                if "Number of Participants" in request.form:
                    NUM_PARTICIPANTS = int(
                        request.form["Number of Participants"])
                    NUM_TEAMS = round(len(members) / int(NUM_PARTICIPANTS))

                else:
                    NUM_TEAMS = int(request.form["groups"])
                initial_solution = Solution()
            else:
                pass

            for name, average, gender in members:  # create random groups for the beginning
                initial_solution.add_player_to_random_team(
                    Player(name, average, gender))

            for run_num in range(NUM_RUNS):  # creating groups

                if run_num % 1000 == 0:
                    print(
                        "Current best solution with score {:.2f}:\n{}".format(  # print the current score
                            initial_solution.group_score(), initial_solution
                        )
                    )

                solution = copy.deepcopy(
                    initial_solution
                )  # copy the current solution object into new variable to compare after

                for _ in range(random.randint(1, 10)):
                    solution.change_random_player()
                if solution.group_score() < initial_solution.group_score():
                    initial_solution = solution

            print(
                "Best solution found, with solution score {:.2f}:\n{}".format(
                    initial_solution.group_score(), initial_solution
                )
            )
            print("\n\n")

            # output the optimal solution to excel file
            excel_dict = initial_solution.exportXl()
            OptionArr = OptionArray()
            OptionArr.setExcel_Dict(excel_dict, option)
            return render_template(
                "complete.html",
                location=os.path.join(
                    "./static/output", secure_filename(workbookName)),
                excel_dict=OptionArr
            )
        return render_template("signup.html", access="Denied")


@app.after_request
def after_request(response):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return response


##
@app.route("/profile-setting", methods=["GET", "POST"])
def profileSetting():
    errormsg = ErrorMessages()
    if "email" in session:
        if request.method == "POST":
            if request.form['action'] == "update":
                fname = request.form['fname']
                lname = request.form['lname']
                password = request.form['password']
                confirmPassword = request.form['confirmpass']
                organization = request.form['organization']

                if password == confirmPassword:
                    hashedPassword = generate_password_hash(password)
                    try:
                        cur = mysql.connection.cursor()
                        cur.execute("UPDATE users SET fname=%s,lname=%s,password=%s,organization=%s WHERE email=%s",
                                    (fname, lname, hashedPassword, organization, session['email']))
                        msg = Message('Profile Changed',
                                      sender="legendbest123@gmail.com", recipients=[session['email']])
                        msg.body = 'Your profile is updated!'
                        mail.send(msg)

                        errormsg.profileUpdated()
                        mysql.connection.commit()
                        cur.close()
                        session.pop('fname', None)
                        session.pop('lname', None)
                        session.pop('organization', None)
                        session.pop('name', None)
                        session['fname']=fname
                        session['lname']=lname
                        session['organization']=organization
                        session['name']=fname + " " + lname
                        return render_template('profile-setting.html', error=errormsg)
                    except SMTPAuthenticationError:
                        errormsg.serverBusy()
                    except:
                        errormsg.serverBusy()
                    return render_template('profile-setting.html', error=errormsg)
                else:
                    errormsg.passDoesntMatched()
                    return render_template("profile-setting.html", error=errormsg)

        else:
            return render_template("profile-setting.html")
    else:
        errormsg.accessDenied()
        return render_template("signup.html", access=errormsg)

@app.after_request
def after_request(response):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return response

if __name__ == "__main__":
    app.run("127.0.0.1",port=3000, debug=True)
